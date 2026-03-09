"""
推薦リスト（.xlsx）生成モジュール
"""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# 推薦度の色設定
GRADE_COLORS = {
    "◎": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # 薄い緑
    "○": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),  # 薄い青
    "△": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # 薄いオレンジ
    "×": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # 薄い赤
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)


def generate_recommendation_xlsx(
    scored_jobs: list[dict],
    candidate_name: str,
    source: str = "circus",
) -> io.BytesIO:
    """
    精密採点済みの求人リストから推薦用xlsxを生成

    scored_jobs: [
        {
            "company": "HENNGE株式会社",
            "title": "セールス",
            "url": "https://...",  (HITO-Linkは空)
            "grade": "◎",
            "reason": "判定理由",
            "industry": "IT",
            "sub_industry": "SaaS",
        }, ...
    ]

    Returns: BytesIO (xlsxバイナリ)
    """
    wb = Workbook()
    ws = wb.active

    today = datetime.now().strftime("%Y%m%d")
    ws.title = f"推薦リスト_{candidate_name}"

    # ヘッダー
    if source == "hito-link":
        headers = ["No.", "推薦度", "大カテゴリ(業界)", "中カテゴリ(業種)",
                    "企業名", "求人名", "判定理由"]
    else:
        headers = ["No.", "推薦度", "大カテゴリ(業界)", "中カテゴリ(業種)",
                    "企業名", "求人名", "URL", "判定理由"]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # ソート: 推薦度順（◎→○→×）、同じ推薦度内では業界→業種
    grade_order = {"◎": 0, "○": 1, "△": 2, "×": 3}
    sorted_jobs = sorted(
        scored_jobs,
        key=lambda j: (
            grade_order.get(j.get("grade", "○"), 4),
            j.get("industry", ""),
            j.get("sub_industry", ""),
        )
    )

    # データ行
    for row_idx, job in enumerate(sorted_jobs, 2):
        no = row_idx - 1
        grade = job.get("grade", "○")

        if source == "hito-link":
            row_data = [
                no,
                grade,
                job.get("industry", ""),
                job.get("sub_industry", ""),
                job.get("company", ""),
                job.get("title", ""),
                job.get("reason", ""),
            ]
        else:
            row_data = [
                no,
                grade,
                job.get("industry", ""),
                job.get("sub_industry", ""),
                job.get("company", ""),
                job.get("title", ""),
                job.get("url", ""),
                job.get("reason", ""),
            ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            # 推薦度セルに色付け
            if col_idx == 2 and grade in GRADE_COLORS:
                cell.fill = GRADE_COLORS[grade]
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # URL列にハイパーリンク
            if source != "hito-link" and col_idx == 7 and value:
                try:
                    cell.hyperlink = value
                    cell.font = Font(color="0563C1", underline="single")
                except Exception:
                    pass

    # 列幅調整
    if source == "hito-link":
        col_widths = [6, 10, 16, 16, 30, 50, 60]
    else:
        col_widths = [6, 10, 16, 16, 30, 50, 50, 60]

    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # ヘッダー行の高さ
    ws.row_dimensions[1].height = 30

    # フィルター有効化
    ws.auto_filter.ref = ws.dimensions

    # ウィンドウ枠固定（ヘッダー行）
    ws.freeze_panes = "A2"

    # BytesIOに書き出し
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_csv_match_xlsx(
    jobs: list[dict],
    candidate_name: str,
    source: str = "circus",
) -> io.BytesIO:
    """
    Phase1のCSVスキルマッチ結果をxlsx形式で出力

    jobs: [{"company": "...", "title": "...", "match_grade": "◎", "match_reason": "...", ...}]
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"求人一覧_{candidate_name}"

    if source == "hito-link":
        headers = ["No.", "ページ", "企業名", "求人タイトル", "年収", "勤務地",
                    "必須スキル", "マッチ度", "判定理由", "メモ"]
    else:
        headers = ["No.", "ページ", "企業名", "求人タイトル", "求人種別", "年収", "勤務地",
                    "必須スキル", "マッチ度", "判定理由", "メモ"]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for row_idx, job in enumerate(jobs, 2):
        grade = job.get("match_grade", "○")

        if source == "hito-link":
            row_data = [
                row_idx - 1,
                job.get("page", ""),
                job.get("company", ""),
                job.get("title", ""),
                job.get("salary", ""),
                job.get("location", ""),
                job.get("skills", "")[:200],
                grade,
                job.get("match_reason", ""),
                "",
            ]
        else:
            row_data = [
                row_idx - 1,
                job.get("page", ""),
                job.get("company", ""),
                job.get("title", ""),
                job.get("job_type", ""),
                job.get("salary", ""),
                job.get("location", ""),
                job.get("skills", "")[:200],
                grade,
                job.get("match_reason", ""),
                "",
            ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

            # マッチ度列に色付け（ページ列追加分 +1）
            match_col = 8 if source == "hito-link" else 9
            if col_idx == match_col and grade in GRADE_COLORS:
                cell.fill = GRADE_COLORS[grade]
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # 列幅（ページ列追加）
    if source == "hito-link":
        col_widths = [6, 8, 30, 50, 16, 16, 40, 10, 50, 20]
    else:
        col_widths = [6, 8, 30, 50, 14, 16, 16, 40, 10, 50, 20]

    for idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
